from __future__ import annotations

import http.server
import json
import os
import pathlib
import shutil
import socket
import socketserver
import subprocess
import sys
import threading
import time
import tkinter as tk
import tkinter.messagebox as messagebox
import urllib.parse
import urllib.request
import webbrowser
from typing import Callable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


APP_TITLE = "Naver Trend Maker 10 로컬버전"
WINDOW_SIZE = "620x280"
UI_PORT = 32110
API_PORT = 8787
REPO_ROOT = pathlib.Path(r"C:\Users\imda0\Desktop\커서 ai 폴더\naver-trend-maker-10")
API_HEALTH_URL = f"http://127.0.0.1:{API_PORT}/v1/health"
API_BASE_URL = f"http://127.0.0.1:{API_PORT}/v1"
LOG_PATH = REPO_ROOT / "local_api.log"
PUMP_INTERVAL_SECONDS = 6
PUMP_TIMEOUT_SECONDS = 180
BEST_PRODUCTS_EXPORT_INTERVAL_SECONDS = 3
BEST_PRODUCTS_EXPORT_TIMEOUT_SECONDS = 10
BEST_PRODUCTS_EXPORT_URL = f"{API_BASE_URL}/products/best/export"
BEST_PRODUCTS_WORKBOOK_PATH = pathlib.Path.home() / "Desktop" / "Naver Trend Maker 10 베스트상품.xlsx"
EXISTING_API_HEALTH_TIMEOUT_SECONDS = 8
API_WATCHDOG_INTERVAL_SECONDS = 5
API_WATCHDOG_FAILURE_THRESHOLD = 2
BEST_PRODUCTS_WORKBOOK_HEADERS = (
    "전체순위",
    "분석점수",
    "키워드점수",
    "기준월",
    "트렌드키워드",
    "키워드순위",
    "등장월수",
    "수집시각",
    "상태",
    "카테고리",
    "분석카드",
    "분석순위",
    "분석키워드",
    "분석근거",
    "최근점수",
    "변화량",
    "모멘텀",
    "계절지수",
    "추천월",
    "주의월",
    "소스",
    "실패사유",
)


def resolve_site_dir() -> pathlib.Path:
    if getattr(sys, "frozen", False):
        return pathlib.Path(sys._MEIPASS) / "site"  # type: ignore[attr-defined]
    return pathlib.Path(__file__).resolve().parent / "web" / ".next-prod"


def is_port_open(port: int) -> bool:
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
        sock.settimeout(0.4)
        return sock.connect_ex(("127.0.0.1", port)) == 0


def wait_for_api(timeout_seconds: int = 45) -> bool:
    deadline = time.time() + timeout_seconds
    while time.time() < deadline:
        try:
            with urllib.request.urlopen(API_HEALTH_URL, timeout=3) as response:
                if response.status == 200:
                    return True
        except Exception:
            time.sleep(1)
    return False


def is_api_healthy(timeout_seconds: int = 2) -> bool:
    try:
        with urllib.request.urlopen(API_HEALTH_URL, timeout=timeout_seconds) as response:
            return response.status == 200
    except Exception:
        return False


def is_best_products_api_ready(timeout_seconds: int = 2) -> bool:
    try:
        with urllib.request.urlopen(BEST_PRODUCTS_EXPORT_URL, timeout=timeout_seconds) as response:
            return response.status == 200
    except Exception:
        return False


def append_log(message: str) -> None:
    with LOG_PATH.open("a", encoding="utf-8", errors="replace") as log:
        log.write(f"{time.strftime('%Y-%m-%d %H:%M:%S')} {message}\n")


def corepack_command() -> str:
    return shutil.which("corepack.cmd") or shutil.which("corepack") or "corepack"


def run_schema_setup() -> None:
    command = [
        corepack_command(),
        "pnpm",
        "wrangler",
        "d1",
        "execute",
        "naver-trend-maker-db",
        "--local",
        "--file",
        "edge-api/schema.sql",
        "--config",
        "edge-api/wrangler.jsonc",
    ]
    with LOG_PATH.open("a", encoding="utf-8", errors="replace") as log:
        log.write("\n\n=== Apply local D1 schema ===\n")
        result = subprocess.run(command, cwd=REPO_ROOT, stdout=log, stderr=subprocess.STDOUT, text=True)
    if result.returncode != 0:
        raise RuntimeError("로컬 데이터베이스 준비에 실패했습니다.")


def stop_stale_local_api_processes() -> None:
    repo_path = str(REPO_ROOT).replace("'", "''")
    command = [
        "powershell.exe",
        "-NoProfile",
        "-Command",
        (
            "$repo = '" + repo_path + "'; "
            "$targets = Get-CimInstance Win32_Process | Where-Object { "
            "($_.Name -in @('node.exe', 'cmd.exe')) -and $_.CommandLine -and "
            "(($_.CommandLine -like ('*' + $repo + '*edge-api/wrangler.jsonc*')) -or "
            "($_.CommandLine -like '*wrangler*--port*8787*')) "
            "}; "
            "$targets | ForEach-Object { Stop-Process -Id $_.ProcessId -Force -ErrorAction SilentlyContinue }"
        ),
    ]
    subprocess.run(command, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL, check=False)


def start_local_api() -> subprocess.Popen[str] | None:
    if is_port_open(API_PORT):
        existing_api_ready = wait_for_api(EXISTING_API_HEALTH_TIMEOUT_SECONDS) and is_best_products_api_ready()
        if not existing_api_ready:
            append_log("[api] port is open but required API readiness failed; cleaning stale local API processes")
            stop_stale_local_api_processes()
            time.sleep(2)
            if is_port_open(API_PORT):
                raise RuntimeError(
                    f"로컬 API 포트({API_PORT})가 이미 사용 중이지만 Naver Trend Maker API가 응답하지 않습니다. "
                    "해당 포트를 쓰는 프로그램을 종료한 뒤 다시 실행해 주세요."
                )
        else:
            return None

    stop_stale_local_api_processes()
    run_schema_setup()
    command = [
        corepack_command(),
        "pnpm",
        "wrangler",
        "dev",
        "--config",
        "edge-api/wrangler.jsonc",
        "--local",
        "--port",
        str(API_PORT),
        "--inspector-port",
        "0",
    ]
    with LOG_PATH.open("a", encoding="utf-8", errors="replace") as log:
        log.write("\n\n=== Start local API ===\n")
        process = subprocess.Popen(
            command,
            cwd=REPO_ROOT,
            stdout=log,
            stderr=subprocess.STDOUT,
            text=True,
            creationflags=subprocess.CREATE_NEW_PROCESS_GROUP,
        )

    if not wait_for_api():
        stop_process_tree(process)
        raise RuntimeError("로컬 API가 시간 안에 시작되지 않았습니다.")

    return process


def start_api_watchdog(
    stop_event: threading.Event,
    api_process_ref: dict[str, subprocess.Popen[str] | None],
) -> threading.Thread:
    def watchdog_loop() -> None:
        failure_count = 0

        while not stop_event.wait(API_WATCHDOG_INTERVAL_SECONDS):
            process = api_process_ref.get("api_process")
            process_exited = process is not None and process.poll() is not None

            if not process_exited and is_api_healthy():
                failure_count = 0
                continue

            failure_count += 1
            append_log(
                f"[api-watchdog] unhealthy count={failure_count} process_exited={process_exited}"
            )

            if failure_count < API_WATCHDOG_FAILURE_THRESHOLD:
                continue

            try:
                stop_process_tree(process)
                stop_stale_local_api_processes()
                api_process_ref["api_process"] = start_local_api()
                failure_count = 0
                append_log("[api-watchdog] local API restored")
            except Exception as error:
                append_log(f"[api-watchdog] restart failed {error}")

    thread = threading.Thread(target=watchdog_loop, name="local-api-watchdog", daemon=True)
    thread.start()
    return thread


def post_worker_process_next() -> str:
    request = urllib.request.Request(
        f"{API_BASE_URL}/trends/worker/process-next",
        method="POST",
        headers={"Content-Type": "application/json"},
    )
    with urllib.request.urlopen(request, timeout=PUMP_TIMEOUT_SECONDS) as response:
        return response.read(800).decode("utf-8", errors="replace")


def start_collection_pump(stop_event: threading.Event) -> threading.Thread:
    def pump_loop() -> None:
        stop_event.wait(2)

        while not stop_event.is_set():
            try:
                if is_port_open(API_PORT):
                    body = post_worker_process_next()
                    append_log(f"[pump] process-next ok {body}")
            except Exception as error:
                append_log(f"[pump] process-next failed {error}")

            stop_event.wait(PUMP_INTERVAL_SECONDS)

    thread = threading.Thread(target=pump_loop, name="trend-collection-pump", daemon=True)
    thread.start()
    return thread


def fetch_best_products_export() -> dict:
    with urllib.request.urlopen(BEST_PRODUCTS_EXPORT_URL, timeout=BEST_PRODUCTS_EXPORT_TIMEOUT_SECONDS) as response:
        payload = response.read(5_000_000).decode("utf-8", errors="replace")
    return json.loads(payload)


def write_best_products_workbook(workbook_path: pathlib.Path, payload: dict) -> None:
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "best_products"
    sheet.append(BEST_PRODUCTS_WORKBOOK_HEADERS)

    header_fill = PatternFill("solid", fgColor="1F6F5C")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for item in payload.get("items", []):
        sheet.append(
            [
                item.get("globalRank") or "",
                item.get("bestScore") or "",
                item.get("keywordScore") or "",
                item.get("trendPeriod", ""),
                item.get("trendKeyword", ""),
                item.get("trendRank") or "",
                item.get("keywordAppearanceCount") or "",
                item.get("collectedAt", ""),
                item.get("status", ""),
                item.get("categoryName") or _last_category_segment(str(item.get("categoryPath", ""))),
                item.get("analysisCard") or item.get("query", ""),
                item.get("rank") or "",
                item.get("title", ""),
                item.get("analysisRationale", ""),
                item.get("latestScore", ""),
                item.get("delta", ""),
                item.get("momentum", ""),
                item.get("seasonalIndex", ""),
                ", ".join(item.get("recommendedMonths", []) or []),
                ", ".join(item.get("cautionMonths", []) or []),
                item.get("source", ""),
                item.get("failureReason", ""),
            ]
        )

    sheet.freeze_panes = "A2"
    widths = {
        "A": 10,
        "B": 12,
        "C": 12,
        "D": 12,
        "E": 22,
        "F": 10,
        "G": 10,
        "H": 24,
        "I": 12,
        "J": 18,
        "K": 18,
        "L": 12,
        "M": 42,
        "N": 12,
        "O": 18,
        "P": 16,
        "Q": 16,
        "R": 20,
        "S": 44,
        "T": 44,
        "U": 28,
        "V": 42,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    temp_path = workbook_path.with_suffix(".tmp.xlsx")
    workbook.save(temp_path)
    os.replace(temp_path, workbook_path)


def _last_category_segment(category_path: str) -> str:
    parts = [part.strip() for part in category_path.split(">") if part.strip()]
    return parts[-1] if parts else ""


def start_best_products_export_pump(stop_event: threading.Event) -> threading.Thread:
    def pump_loop() -> None:
        last_signature = ""
        stop_event.wait(3)

        while not stop_event.is_set():
            try:
                if is_port_open(API_PORT):
                    payload = fetch_best_products_export()
                    signature = json.dumps(payload.get("items", []), ensure_ascii=False, sort_keys=True)
                    if payload.get("ok") and signature != last_signature:
                        write_best_products_workbook(BEST_PRODUCTS_WORKBOOK_PATH, payload)
                        last_signature = signature
                        append_log(f"[best-products] workbook updated {BEST_PRODUCTS_WORKBOOK_PATH}")
            except Exception as error:
                append_log(f"[best-products] workbook update failed {error}")

            stop_event.wait(BEST_PRODUCTS_EXPORT_INTERVAL_SECONDS)

    thread = threading.Thread(target=pump_loop, name="best-products-export-pump", daemon=True)
    thread.start()
    return thread


class ExportRequestHandler(http.server.SimpleHTTPRequestHandler):
    def __init__(self, *args, directory: str, **kwargs):
        super().__init__(*args, directory=directory, **kwargs)

    def log_message(self, format: str, *args) -> None:
        return

    def do_GET(self) -> None:
        self.path = self.rewrite_path(self.path)
        super().do_GET()

    def do_HEAD(self) -> None:
        self.path = self.rewrite_path(self.path)
        super().do_HEAD()

    def rewrite_path(self, raw_path: str) -> str:
        parts = urllib.parse.urlsplit(raw_path)
        clean_path = urllib.parse.unquote(parts.path)

        if not clean_path or clean_path == "/":
            return raw_path

        local_root = pathlib.Path(self.directory)
        relative_path = clean_path.lstrip("/")
        direct_path = local_root / relative_path
        html_path = local_root / f"{relative_path}.html"
        index_path = local_root / relative_path / "index.html"

        if direct_path.exists():
            return raw_path

        if html_path.exists():
            updated_path = f"{clean_path}.html"
            return urllib.parse.urlunsplit(("", "", updated_path, parts.query, parts.fragment))

        if index_path.exists():
            updated_path = f"{clean_path.rstrip('/')}/"
            return urllib.parse.urlunsplit(("", "", updated_path, parts.query, parts.fragment))

        return raw_path


class ThreadedHTTPServer(socketserver.ThreadingMixIn, http.server.HTTPServer):
    daemon_threads = True


def start_ui_server(site_dir: pathlib.Path) -> tuple[ThreadedHTTPServer, int]:
    handler = lambda *args, **kwargs: ExportRequestHandler(*args, directory=str(site_dir), **kwargs)
    server = ThreadedHTTPServer(("127.0.0.1", UI_PORT), handler)
    thread = threading.Thread(target=server.serve_forever, daemon=True)
    thread.start()
    return server, int(server.server_port)


def stop_process_tree(process: subprocess.Popen[str] | None) -> None:
    if process is None or process.poll() is not None:
        return
    subprocess.run(
        ["taskkill", "/PID", str(process.pid), "/T", "/F"],
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        check=False,
    )


def open_browser(url: str) -> None:
    if not webbrowser.open(url):
        messagebox.showinfo(APP_TITLE, f"브라우저에서 이 주소를 직접 여세요:\n{url}")


def build_shutdown_handler(
    root: tk.Tk,
    ui_server: ThreadedHTTPServer,
    api_process_ref: dict[str, subprocess.Popen[str] | None],
    pump_stop_event: threading.Event,
) -> Callable[[], None]:
    def shutdown_app() -> None:
        pump_stop_event.set()
        ui_server.shutdown()
        ui_server.server_close()
        stop_process_tree(api_process_ref.get("api_process"))
        root.destroy()

    return shutdown_app


def build_window(
    url: str,
    ui_server: ThreadedHTTPServer,
    api_process_ref: dict[str, subprocess.Popen[str] | None],
    pump_stop_event: threading.Event,
) -> tk.Tk:
    root = tk.Tk()
    root.title(APP_TITLE)
    root.geometry(WINDOW_SIZE)
    root.resizable(False, False)

    frame = tk.Frame(root, padx=20, pady=20)
    frame.pack(fill="both", expand=True)

    header = tk.Label(
        frame,
        text="로컬 API와 관리자 화면을 실행했습니다.",
        font=("Malgun Gothic", 14, "bold"),
        anchor="w",
        justify="left",
    )
    header.pack(fill="x")

    lines = [
        "Cloudflare 가입 없이 이 PC 안에서만 작동합니다.",
        "자동 순회 중 브라우저 탭을 닫으면 현재 취합을 취소합니다.",
        "이 창을 닫으면 로컬 서버와 백그라운드 취합도 같이 종료됩니다.",
        "",
        f"관리자 화면: {url}",
        f"API 주소: {API_BASE_URL}",
        f"베스트상품 엑셀: {BEST_PRODUCTS_WORKBOOK_PATH}",
    ]
    body = tk.Label(frame, text="\n".join(lines), font=("Malgun Gothic", 10), anchor="w", justify="left")
    body.pack(fill="x", pady=(14, 18))

    button_row = tk.Frame(frame)
    button_row.pack(fill="x")

    shutdown_app = build_shutdown_handler(root, ui_server, api_process_ref, pump_stop_event)

    tk.Button(button_row, text="브라우저 다시 열기", width=18, command=lambda: open_browser(url)).pack(side="left")
    tk.Button(button_row, text="종료", width=12, command=shutdown_app).pack(side="right")

    root.protocol("WM_DELETE_WINDOW", shutdown_app)
    return root


def main() -> int:
    site_dir = resolve_site_dir()
    if not site_dir.exists():
        messagebox.showerror(APP_TITLE, f"화면 파일을 찾지 못했습니다.\n\n{site_dir}")
        return 1

    try:
        api_process = start_local_api()
    except Exception as error:
        messagebox.showerror(APP_TITLE, f"{error}\n\n로그 파일:\n{LOG_PATH}")
        return 1

    try:
        ui_server, port = start_ui_server(site_dir)
    except OSError:
        stop_process_tree(api_process)
        messagebox.showerror(
            APP_TITLE,
            "앱 전용 로컬 포트(32110)를 열지 못했습니다.\n\n"
            "이미 실행 중인 Naver Trend Maker 창을 닫고 다시 실행해 주세요.",
        )
        return 1

    url = f"http://127.0.0.1:{port}/sourcing/admin.html"
    api_process_ref = {"api_process": api_process}
    pump_stop_event = threading.Event()
    start_api_watchdog(pump_stop_event, api_process_ref)
    start_collection_pump(pump_stop_event)
    start_best_products_export_pump(pump_stop_event)
    open_browser(url)
    window = build_window(url, ui_server, api_process_ref, pump_stop_event)
    window.mainloop()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
