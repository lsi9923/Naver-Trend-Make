from __future__ import annotations

import importlib.util
import pathlib
import tempfile
import unittest

from openpyxl import load_workbook


PROJECT_ROOT = pathlib.Path(__file__).resolve().parents[1]
MODULE_PATH = PROJECT_ROOT / "local_app_launcher.py"


def load_launcher_module():
    spec = importlib.util.spec_from_file_location("local_app_launcher_under_test", MODULE_PATH)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Could not load {MODULE_PATH}")

    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


class BestProductsExcelExportTests(unittest.TestCase):
    def test_workbook_is_rewritten_with_current_accumulated_best_products(self) -> None:
        launcher = load_launcher_module()

        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = pathlib.Path(temp_dir) / "best-products.xlsx"

            launcher.write_best_products_workbook(
                workbook_path,
                {
                    "ok": True,
                    "items": [
                        self.analysis_candidate("니트", 1, "울 니트 A"),
                        self.analysis_candidate("니트", 2, "캐시미어 니트 B"),
                    ],
                },
            )
            launcher.write_best_products_workbook(
                workbook_path,
                {
                    "ok": True,
                    "items": [
                        self.analysis_candidate("니트", 1, "업데이트 니트 A"),
                        self.analysis_candidate("니트", 2, "캐시미어 니트 B"),
                        self.analysis_candidate("원피스", 1, "플리츠 원피스 A"),
                        self.analysis_candidate("원피스", 2, "셔츠 원피스 B"),
                    ],
                },
            )

            workbook = load_workbook(workbook_path)
            sheet = workbook["best_products"]
            rows = list(sheet.iter_rows(values_only=True))

        self.assertEqual(
            rows[0][:12],
            (
                "전체순위",
                "분석점수",
                "키워드점수",
                "기준월",
                "트렌드키워드",
                "키워드순위",
                "등장월수",
                "수집시각",
                "상태",
                "카테고리",
                "분석카드",
                "분석순위",
            ),
        )
        self.assertEqual([row[9] for row in rows[1:]], ["니트", "니트", "원피스", "원피스"])
        self.assertEqual([row[12] for row in rows[1:]], ["업데이트 니트 A", "캐시미어 니트 B", "플리츠 원피스 A", "셔츠 원피스 B"])

    def test_workbook_keeps_failure_rows_visible_without_crashing(self) -> None:
        launcher = load_launcher_module()

        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = pathlib.Path(temp_dir) / "best-products.xlsx"
            launcher.write_best_products_workbook(
                workbook_path,
                {
                    "ok": True,
                    "items": [
                        {
                            "collectedAt": "2026-06-08T00:00:00.000Z",
                            "status": "failed",
                            "categoryPath": "패션의류 > 여성의류 > 니트",
                            "categoryName": "니트",
                            "query": "steady",
                            "trendPeriod": "2026-05",
                            "trendKeyword": "여성니트",
                            "trendRank": 1,
                            "keywordScore": 1200,
                            "keywordAppearanceCount": 6,
                            "rank": 0,
                            "title": "",
                            "analysisCard": "꾸준히 스테디하게 판매하기 좋은 키워드",
                            "analysisRationale": "",
                            "latestScore": 0,
                            "delta": 0,
                            "momentum": 0,
                            "seasonalIndex": 0,
                            "recommendedMonths": [],
                            "cautionMonths": [],
                            "image": "",
                            "lowPrice": None,
                            "mallName": "",
                            "brand": "",
                            "maker": "",
                            "productId": "",
                            "source": "naver-shopping-insight:trend-analysis",
                            "failureReason": "TREND_ANALYSIS_KEYWORDS_MISSING",
                        }
                    ],
                },
            )

            workbook = load_workbook(workbook_path)
            sheet = workbook["best_products"]
            rows = list(sheet.iter_rows(values_only=True))

        self.assertEqual(rows[1][8], "failed")
        self.assertEqual(rows[1][9], "니트")
        self.assertEqual(rows[1][-1], "TREND_ANALYSIS_KEYWORDS_MISSING")

    @staticmethod
    def analysis_candidate(category_name: str, rank: int, title: str) -> dict[str, object]:
        return {
            "collectedAt": "2026-06-08T00:00:00.000Z",
            "globalRank": rank,
            "bestScore": 1000 - rank,
            "status": "collected",
            "categoryPath": f"패션의류 > 여성의류 > {category_name}",
            "categoryName": category_name,
            "query": "steady",
            "trendPeriod": "2026-05",
            "trendKeyword": f"{category_name} 인기",
            "trendRank": rank,
            "keywordScore": 1200 - rank,
            "keywordAppearanceCount": 5,
            "rank": rank,
            "title": title,
            "analysisCard": "꾸준히 스테디하게 판매하기 좋은 키워드",
            "analysisRationale": f"{category_name} 분석 근거",
            "latestScore": 18,
            "delta": 2.5,
            "momentum": 1.2,
            "seasonalIndex": 1.4,
            "recommendedMonths": ["5월", "6월"],
            "cautionMonths": ["1월"],
            "link": "",
            "image": "",
            "lowPrice": 10000 + rank,
            "mallName": "테스트몰",
            "brand": "",
            "maker": "",
            "productId": f"{category_name}-{rank}",
            "source": "naver-shopping-insight:trend-analysis",
            "failureReason": "",
        }


if __name__ == "__main__":
    unittest.main()
